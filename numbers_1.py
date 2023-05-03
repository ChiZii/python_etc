import os, pathlib, threading, platform

from tkinter import Button, messagebox
import openpyxl
from bs4 import BeautifulSoup
import requests
import matplotlib.pyplot as plt

import pymongo
import datetime

class App:
  def __init__(self) -> None:
    t = datetime.datetime.now()
    self.today = '{:04d}{:02d}{:02d}'.format(t.year, t.month, t.day)
    self.now = '{:02d}{:02d}'.format(t.hour, t.minute)
    self.numberFilePath = os.path.join(pathlib.Path.home(), 'Downloads', 'LTnumbers.xlsx')
    self.uri = ''

    self.createFile()

  def algorithm(self, wb: openpyxl.Workbook):
    ws_num = wb.worksheets[0]
    ws = wb.worksheets[1]
    
    temp = []
    for i,v in enumerate(ws_num.rows):
        if i == 0: continue
        obj = {
        '_id':int(v[0].value),
        '1st':int(v[1].value),
        '2nd':int(v[2].value),
        
        '3rd':int(v[3].value),
        '4th':int(v[4].value),
        '5th':int(v[5].value),
        '6th':int(v[6].value),
        'Bns':int(v[7].value),
        }
        temp.append(obj)
    for no in range(1,8):
      pos=[]
      eqr=[]
      neg=[]
      fac = []
      for i, v in enumerate(temp):
          tempFac = []
          if i == 0: continue
          tempFac.append(v['_id'])
          tempFac.append(temp[i-1]['1st'] - v['1st'])
          tempFac.append(temp[i-1]['2nd'] - v['2nd'])
          tempFac.append(temp[i-1]['3rd'] - v['3rd'])
          tempFac.append(temp[i-1]['4th'] - v['4th'])
          tempFac.append(temp[i-1]['5th'] - v['5th'])
          tempFac.append(temp[i-1]['6th'] - v['6th'])
          tempFac.append(temp[i-1]['Bns'] - v['Bns'])
          fac.append(tempFac)

          position = tempFac[no]
          if position > 0:
              pos.append(tempFac[no])
          elif position < 0:
              neg.append(tempFac[no])
          elif position == 0:
              eqr.append(tempFac[no])

      pos.sort()
      neg.sort()
      posDup = list(dict.fromkeys(pos))
      negDup = list(dict.fromkeys(neg))
      print(no, max(pos), len(eqr),min(neg))
      print(posDup)
      print(negDup)
    wb.save(self.numberFilePath)
    # if platform.system() == 'windows':
    #   os.startfile(self.numberFilePath)
    # else:
    #   os.system("open {}".format(self.numberFilePath))

  def createFile(self):
    t = datetime.datetime.now()
    today = '{:04d}{:02d}{:02d}'.format(t.year, t.month, t.day)
    now = '{:02d}{:02d}'.format(t.hour, t.minute)

    dvg = 'normal'

    if os.path.isfile(self.numberFilePath):
      wb_num = openpyxl.load_workbook(self.numberFilePath)
      ws = wb_num.worksheets[0]
      createDate = ws.cell(row=1, column=1).value
      thatDay = datetime.date.fromisoformat('{}-{}-{}'.format(createDate[0:4], createDate[4:6], createDate[6:8]))
      saturday = thatDay - datetime.timedelta(days=thatDay.weekday()+2)
    
      if ((t.date() - saturday).days == 0 and int(now) > 2100) or (t.date() - saturday).days > 7:
          dvg = 'newNumber'

    else:
      dvg = 'newFile'


    if dvg != 'normal':
      client = pymongo.MongoClient(self.uri)
      try:
          client.admin.command('ping')
          print("Load MongoDB for new file!")
      except Exception as e:
          print(e)
          return
      
      db = client['numbers']
      colNum = db['numbers']
      colFac = db['factors']
      
      # Insert New Number
      try:
          r = requests.get('https://dhlottery.co.kr/common.do?method=main')
      except:
         print('Fail, Try again')
      soup = BeautifulSoup(r.text, 'html.parser')

      tempObj = {
      '_id':int(soup.find(id='lottoDrwNo').text),
      '1st':int(soup.find(id='drwtNo1').text),
      '2nd':int(soup.find(id='drwtNo2').text),
      '3rd':int(soup.find(id='drwtNo3').text),
      '4th':int(soup.find(id='drwtNo4').text),
      '5th':int(soup.find(id='drwtNo5').text),
      '6th':int(soup.find(id='drwtNo6').text),
      'Bns':int(soup.find(id='bnusNo').text)}


      recentNo = colNum.find().sort('_id',-1).limit(1)[0]['_id']
      if tempObj['_id'] == recentNo + 1:
          colNum.insert_one(tempObj)
      elif tempObj['_id'] > recentNo + 1:
          messagebox.showwarning(title='Too Many Skip', message='{} numbers are missing!'.format(tempObj['_id']-recentNo))
          return

      match dvg:
          case 'newFile':
              wb_num = openpyxl.Workbook()
      ws_create = wb_num.worksheets[0]
      ws_create.cell(row=1, column=1).value = today + now
      ws_create.cell(row=1, column=2).value = '1st'
      ws_create.cell(row=1, column=3).value = '2nd'
      ws_create.cell(row=1, column=4).value = '3rd'
      ws_create.cell(row=1, column=5).value = '4th'
      ws_create.cell(row=1, column=6).value = '5th'
      ws_create.cell(row=1, column=7).value = '6th'
      ws_create.cell(row=1, column=8).value = 'Bns'
      idx = 2
      for i in colNum.find().sort('_id'):
        ws_create.cell(row=idx, column=1).value = i['_id']
        ws_create.cell(row=idx, column=2).value = i['1st']
        ws_create.cell(row=idx, column=3).value = i['2nd']
        ws_create.cell(row=idx, column=4).value = i['3rd']
        ws_create.cell(row=idx, column=5).value = i['4th']
        ws_create.cell(row=idx, column=6).value = i['5th']
        ws_create.cell(row=idx, column=7).value = i['6th']
        ws_create.cell(row=idx, column=8).value = i['Bns']
        idx+=1
      wb_num.save(self.numberFilePath)

      client.close()

    if len(wb_num.sheetnames) == 1:
      wb_num.create_sheet('factor')

    ws_factor = wb_num.worksheets[1]
    wb_num.active = ws_factor
    wb_num.save(self.numberFilePath)
    
    self.algorithm(wb_num)

App()